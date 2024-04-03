from urllib import response
from django.shortcuts import render,HttpResponse,redirect
from django.contrib.auth.models import User
from django.contrib.auth import authenticate,login,logout
from django.contrib.auth.decorators import login_required

from app1.models import contactEnquiry
from app1.forms import EmployeeForm
from app1 import views
from .filters import OrderFilters

from django.core.paginator import Paginator,EmptyPage,PageNotAnInteger
from openpyxl import Workbook, workbook
from openpyxl.styles import *
import decimal
from app1.models import CountryGDP
from tablib import Dataset
from .resources import PersonResourc

# Create your views here.
@login_required(login_url='login')
def HomePage(request):
    contactdata=contactEnquiry.objects.all()
    
    filters=OrderFilters(request.GET,queryset=contactdata)
    data={'contactdata':contactdata,'filters':filters}
    return render (request,'home.html',data)

def SignupPage(request):
    if request.method=='POST':
        uname=request.POST.get('username')
        email=request.POST.get('email')
        pass1=request.POST.get('password1')
        pass2=request.POST.get('password2')

        if pass1!=pass2:
            return HttpResponse("Your password and confrom password are not Same!!")
        else:

            my_user=User.objects.create_user(uname,email,pass1)
            my_user.save()
            return redirect('login')
        

    return render (request,'signup.html')

def LoginPage(request):
    if request.method=='POST':
        username=request.POST.get('username')
        pass1=request.POST.get('pass')

        user=authenticate(request,username=username,password=pass1)

        if user is not None:
            login(request,user)
            return redirect('home')
        else:
            return HttpResponse ("Username or Password is incorrect!!!")

    return render (request,'login.html')

def LogoutPage(request):
    logout(request)
    return redirect('login')

@login_required(login_url='login')

def saveEnquiry(request):
    if request.method=='POST':
        name=request.POST.get('name')
        email=request.POST.get('email')
        password=request.POST.get('password')
        en=contactEnquiry(name=name,email=email,password=password)
        en.save()
        return redirect('home')
    return render(request,'contact.html')


@login_required(login_url='login')
def contactPage(request):
    
    return redirect('saveEnquiry')

@login_required(login_url='login')
def UpdateBlog(request):
    contactdata=contactEnquiry.objects.all()
    data={'contactdata':contactdata}
    return render(request,'update.html',data)

@login_required(login_url='login')

def EditBlog(request,id):

    contactdata=contactEnquiry.objects.get(id=id)

    return render(request,'edit.html',{'contactdata':contactdata})

@login_required(login_url='login')
def up(request,id):
    
    contactdata=contactEnquiry.objects.get(id=id)
    form=EmployeeForm(request.POST,instance=contactdata)
    if form.is_valid():
        form.save()
        return redirect('/update')

    return render(request,'edit.html',{'contactdata':contactdata})
@login_required(login_url='login')

def deleteData(request,id):
    
    contactdata=contactEnquiry.objects.get(id=id)
    contactdata.delete()
    
    return redirect('/update')

@login_required(login_url='login')
def searchEmp(request):
    
    contactdata=contactEnquiry.objects.all()

    filters=OrderFilters(request.GET,queryset=contactdata)

   
    
    return render(request,'searchEmployee.html',{'filters':filters})   
@login_required(login_url='login')
def searchBar(request):
    if request.method=='GET':
        query=request.GET.get('query')
        if query:
            data=contactEnquiry.objects.filter(name__icontains=query)
            return render(request,'home.html',{'data':data})
        else:
            print('No Information about it')
            return render(request,'home.html',{})
        
def is_valid_queryparam(param):
    return param != '' and param is not None


def countries_gdp_list(request):
    qs = CountryGDP.objects.order_by('name')

    name = request.GET.get('name')
    year = request.GET.get('year')

    request.session['name'] = name
    request.session['year'] = year

    if is_valid_queryparam(name):
        qs = qs.filter(name__icontains=name)

    if is_valid_queryparam(year):
        qs = qs.filter(year=year)

    page = request.GET.get('page', 1)
    paginator = Paginator(qs, 30)

    try:
        qs = paginator.page(page)
    except PageNotAnInteger:
        qs = paginator.page(1)
    except EmptyPage:
        qs = paginator.page(paginator.num_pages)

    context = {
        'countries_list': qs,
        'name': name,
        'year':year,
    }
    return render(request, "countries_list.html", context)




def countries_gdp_excel(request):

   
    
    qs = CountryGDP.objects.order_by('name')

    if 'name' in request.session:
        name = request.session['name']
    else:
        name = None

    if 'year' in request.session:
        year = request.session['year']
    else:
        year = None

    if is_valid_queryparam(name):
        qs = qs.filter(name__icontains=name)

    if is_valid_queryparam(year):
        qs = qs.filter(year=year)

    if year is None or year == '':
        year = "2019 - 2021"
    else:
        year = year

    if name is None or name == '':
        name = "All Countries"
    else:
        name = name

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',)
    response['Content-Disposition'] = 'attachment; filename="' + 'Countries GDP List' +'.xlsx"'
    workbook = Workbook()

    worksheet = workbook.active

    worksheet.merge_cells('A1:D1')
    worksheet.merge_cells('A2:D2')
    first_cell = worksheet['A1']
    first_cell.value = "Countries GDP List" + " " + year
    first_cell.fill = PatternFill("solid", fgColor="246ba1")
    first_cell.font  = Font(bold=True, color="F7F6FA")
    first_cell.alignment = Alignment(horizontal="center", vertical="center")

    second_cell = worksheet['A2']
    second_cell.value = name
    second_cell.font  = Font(bold=True, color="246ba1")
    second_cell.alignment = Alignment(horizontal="center", vertical="center")

    worksheet.title = 'Countries GDP List' + " " + year

    # Define the titles for columns
    columns = ['Country Name','Country Code','Year', 'Value in USD']
    row_num = 3

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.fill = PatternFill("solid", fgColor="50C878")
        cell.font  = Font(bold=True, color="F7F6FA")
        third_cell = worksheet['D3']
        third_cell.alignment = Alignment(horizontal="right")

    for countries in qs:
        row_num += 1

        # Define the data for each cell in the row
        row = [countries.name,countries.code,countries.year,countries.value]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
         cell = worksheet.cell(row=row_num, column=col_num)
         cell.value = cell_value
         if isinstance(cell_value, decimal.Decimal):
             cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1

	        

    workbook.save(response)
    return response

def savecountries(request):
    if request.method=='POST':
        name=request.POST.get('name')
        code=request.POST.get('code')
        year=request.POST.get('year')
        value=request.POST.get('value')
        en=CountryGDP(name=name,code=code,value=value,year=year)
        en.save()
        return redirect('countries_gdp_list')
    return render(request,'addcountries.html')

def importExcel(request):
    if request.method=='POST':
       
        dataset=Dataset()
        new_person=request.FILES['my_file']
        imported_data=dataset.load(new_person.read(),format='xlsx')
        for data in imported_data:
            value=CountryGDP(data[0],
                             data[1],
                             data[2],
                             data[3],
                             data[4]
                             )
            value.save()
        return redirect('countries_gdp_list')

    return render(request,'form.html')